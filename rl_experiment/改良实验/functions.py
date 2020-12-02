import random
import networkx as nx
import numpy as np
from Agent import GetInformation
from QTable import QLearningTable
import Constant
import pandas as pd
from openpyxl import load_workbook
import readyaml



def net_work(nodelist, edgelist):
    """获得路网"""
    G = nx.Graph()          # 创建空的无向图
    # 增加节点
    G.add_node(1)                # 每次增加一个节点
    G.add_nodes_from(nodelist)  # iterable container
    #print(list(G.nodes))    # 获取所有节点 [1, 3, 5, 8, 6, 2]
    # 增加edges
    G.add_edges_from(edgelist)
    #nx.draw(G, with_labels=True)
    #plt.show()
    #print(list(G.adj[1]))
    return G


#定义节点的初始状态
def init_NS(node):
    if random.random() > 0.5:
        state=1
    else:
        state=0
    return state


def get_action(action):
    if action == 0:  # 节点i的动作
        z = np.mat(([1], [0]))
    else:
        z = np.mat(([0], [1]))
    return z


"""#初始化每个路口的Agent和Qtable"""


def inital_agent_qtable(the_NET):
    step = 0
    for i in the_NET.nodes():
        if step < 6:
            step += 1
            the_cross = GetInformation(readyaml.id_list[i], readyaml.all_edge_list[i])
            Constant.get_cross_information_list.append(the_cross)
            the_real_Qtable = QLearningTable([0, 1])
            Constant.Real_Q_table_list.append(the_real_Qtable)

            the_net_Qtable = QLearningTable([0, 1])
            Constant.Net_Q_table_list.append(the_net_Qtable)
            one_nodestate = init_NS(readyaml.node_list[i])
            Constant.Node_States_list.append(one_nodestate)
        else:
            break
    return Constant.Real_Q_table_list, Constant.Node_States_list, Constant.get_cross_information_list, Constant.Net_Q_table_list


"""RL take action and get next observation and reward
    真实路网的state：排队长度
    博弈节点根据强化学习选动作
    计算节点i的动作对应的向量"""


def take_state_action_zx(i, newNStates, cross_list, Q_table_list):
    state = cross_list[i].get_current_state()
    agent_action_chosen = Q_table_list[i].select_action(newNStates[i], readyaml.EPSILON)
    zx = get_action(agent_action_chosen)
    return state, agent_action_chosen, zx    #可能会出问题，因为return的值可能为NONE，为此需要提前赋值或使用else


#获得邻居的Q值之和节点i的奖励
def get_nbr_q_i_reward(Net_Q_table_list, get_cross_information_list, Real_Q_table_list, zx, newNStates, j, reward, nbr_Q_list):
    nbr_action = Net_Q_table_list[j].select_action(newNStates[j], readyaml.EPSILON)           # 获取博弈节点邻居的动作
    one_nbr_queque_count = get_cross_information_list[j].get_current_state()  #求邻居的Qtable中GetMax的变量state，即排队长度
    one_nbr_Q = Real_Q_table_list[j].get_max_q(one_nbr_queque_count)*nbr_action      #邻居的q值
    nbr_Q_list.append(one_nbr_Q)
    zy = get_action(nbr_action)#计算节点i的邻居j的动作对应的向量
    reward = reward + zx.T * readyaml.Amat * zy #计算节点i的奖励
    return Constant.nbr_Q_list, reward


def do_one_action(Real_Q_table_list, queue_count, get_cross_information_list, i):
    cross_action_chosen = Real_Q_table_list[i].select_action(queue_count, readyaml.EPSILON)#a
    get_cross_information_list[i].do_action(cross_action_chosen)#执行动作
    return cross_action_chosen


def save_real_q_table(step, Real_Q_table_list, i):
    name = ("F:/研一实验/实验二/RealQTable/RealQTable" + str(i) + ".xlsx")
    writer = pd.ExcelWriter(name, engine='openpyxl')
    book = load_workbook(writer.path)
    writer.book = book
    sheetname = step
    pd.DataFrame(Real_Q_table_list[i].q_table).to_excel(excel_writer=writer, sheet_name=str(sheetname))
    writer.save()
    writer.close()


def save_net_q_table(step, Net_Q_table_list, i):
    name = ("F:/研一实验/实验二/NetQTable/RealQTable" + str(i) + ".xlsx")
    writer = pd.ExcelWriter(name, engine='openpyxl')
    book = load_workbook(writer.path)
    writer.book = book
    sheetname = step
    pd.DataFrame(Net_Q_table_list[i].q_table).to_excel(excel_writer=writer, sheet_name=str(sheetname))
    writer.save()
    writer.close()


def append_in_the_list(list0, list1, list2, element0, element1, element2):
    list0.append(element0)
    list1.append(element1)
    list2.append(element2)
    return list0, list1, list2


def save_csv(number, input_data):
    for i in range(number):
        name = ["queue", "action", "reward", "q_rl", "q_ngl"]
        test = pd.DataFrame(columns=name, data=input_data[i])
        test.to_csv("F:/改良实验/realqtable" + str(i) + ".csv")


